from __future__ import annotations

from io import BytesIO
from typing import Iterable

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


YELLOW_FILL = PatternFill(fill_type="solid", fgColor="FFF2CC")
RED_FILL = PatternFill(fill_type="solid", fgColor="F4CCCC")
HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E78")


def build_excel(items_df: pd.DataFrame, summary_df: pd.DataFrame) -> BytesIO:
    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        items_df.to_excel(writer, sheet_name="Itens", index=False)
        summary_df.to_excel(writer, sheet_name="Resumo", index=False)

        _format_sheet(writer.book["Itens"], items_df)
        _format_sheet(writer.book["Resumo"], summary_df)
        _highlight_rows(
            writer.book["Itens"],
            items_df,
            status_column="status_conferencia",
            status_values={"CONFERIR"},
            fill=YELLOW_FILL,
        )
        _highlight_rows(
            writer.book["Resumo"],
            summary_df,
            status_column="status_pdf",
            status_values={"DIVERGENTE", "ERRO"},
            fill=RED_FILL,
        )
        _apply_number_formats(writer.book["Itens"], items_df)

    output.seek(0)
    return output


def generate_excel(items_df: pd.DataFrame, summary_df: pd.DataFrame) -> BytesIO:
    return build_excel(items_df, summary_df)


def _format_sheet(ws, df: pd.DataFrame) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for column_index, column_name in enumerate(df.columns, start=1):
        letter = get_column_letter(column_index)
        values = [str(column_name)]
        values.extend("" if value is None else str(value) for value in df[column_name].head(500))
        max_length = max((len(value) for value in values), default=12)
        ws.column_dimensions[letter].width = min(max(max_length + 2, 12), 60)


def _highlight_rows(ws, df: pd.DataFrame, *, status_column: str, status_values: Iterable[str], fill: PatternFill) -> None:
    if status_column not in df.columns:
        return

    status_index = list(df.columns).index(status_column) + 1
    wanted = set(status_values)

    for row_number in range(2, ws.max_row + 1):
        if ws.cell(row=row_number, column=status_index).value in wanted:
            for column_number in range(1, ws.max_column + 1):
                ws.cell(row=row_number, column=column_number).fill = fill


def _apply_number_formats(ws, df: pd.DataFrame) -> None:
    number_formats = {
        "peso_g": "0,00",
        "valor_grama_classificacao": "0,00",
        "valor_base": "R$ #.##0,00",
        "valor_unitario": "R$ #.##0,00",
        "valor_total": "R$ #.##0,00",
        "percentual": "0,00",
        "confianca_extracao": "0",
    }

    for column_name, number_format in number_formats.items():
        if column_name not in df.columns:
            continue
        column_index = list(df.columns).index(column_name) + 1
        for row_number in range(2, ws.max_row + 1):
            ws.cell(row=row_number, column=column_index).number_format = number_format
